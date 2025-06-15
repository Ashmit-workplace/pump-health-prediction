import os
import pandas as pd
import numpy as np
from scipy.signal import savgol_filter
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import numbers


def process_and_embed(filepath):
    try:
        df = pd.read_excel(filepath)
        if 'datetime' not in df.columns:
            print(f"Skipping {filepath} - no 'datetime' column found.")
            return

        df['datetime'] = pd.to_datetime(df['datetime'])

        for axis in ['x_mps2', 'y_mps2', 'z_mps2']:
            if axis not in df.columns:
                print(f"Skipping {filepath} - missing {axis} column.")
                return

        # Apply smoothing and residual
        window_length = 11 if len(df) >= 11 else (len(df) // 2 * 2 + 1)
        for axis in ['x', 'y', 'z']:
            df[f'{axis}_smooth'] = savgol_filter(df[f'{axis}_mps2'], window_length, polyorder=2)
            df[f'{axis}_resid'] = df[f'{axis}_mps2'] - df[f'{axis}_smooth']

        # Noise summary
        noise_summary = {}
        for axis in ['x', 'y', 'z']:
            resid_std = df[f'{axis}_resid'].std()
            orig_std = df[f'{axis}_mps2'].std()
            ratio = resid_std / orig_std if orig_std != 0 else 0
            if ratio < 0.05:
                noise_summary[axis] = "Very smooth"
            elif ratio < 0.15:
                noise_summary[axis] = "Slight noise"
            else:
                noise_summary[axis] = "High noise"

        # Save to new Excel file
        out_path = filepath.replace(".xlsx", "_denoised.xlsx")
        df_to_save = df[['datetime', 'x_mps2', 'x_smooth', 'y_mps2', 'y_smooth', 'z_mps2', 'z_smooth']]
        df_to_save.to_excel(out_path, sheet_name='Denoised Data', index=False)

        # Embed chart + format timestamp
        wb = load_workbook(out_path)
        ws = wb['Denoised Data']

        for cell in ws['A'][1:]:  # skip header
            cell.number_format = 'yyyy-mm-dd hh:mm:ss.000'

        for col_start, axis in zip([2, 4, 6], ['x', 'y', 'z']):
            chart = LineChart()
            chart.title = f"{axis.upper()} Axis - Original vs Smoothed"
            chart.y_axis.title = 'm/s²'
            chart.x_axis.title = 'Sample Index'
            data = Reference(ws, min_col=col_start, max_col=col_start + 1, min_row=1, max_row=min(200, len(df)) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.height = 8
            chart.width = 20
            ws.add_chart(chart, f"H{col_start * 10}")

        # Write noise summary
        row = 5
        for axis, note in noise_summary.items():
            ws[f"K{row}"] = f"{axis.upper()} axis: {note}"
            row += 1

        wb.save(out_path)
        print(f"✅ Processed and saved: {out_path}")

    except Exception as e:
        print(f"❌ Error processing {filepath}: {e}")


def recursive_process(root_folder):
    for dirpath, _, filenames in os.walk(root_folder):
        for file in filenames:
            if file.lower().endswith('.xlsx') and not file.endswith('_denoised.xlsx'):
                full_path = os.path.join(dirpath, file)
                process_and_embed(full_path)


# === USAGE ===
# Replace this with your actual top-level folder
folder_to_process = r"D:\extracted data from JSON file ISI\extracted data\denoised_data\sensor_data_cleaned"
recursive_process(folder_to_process)


