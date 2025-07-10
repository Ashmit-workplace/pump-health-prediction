import os
import pandas as pd
import numpy as np
from scipy.fft import fft, fftfreq
from scipy.signal import detrend
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === CONFIGURATION ===
bands = [(0, 1), (1, 3), (3, 5), (5, 10)]  # Only up to 10 Hz
axes = ['x_mps2', 'y_mps2', 'z_mps2']

def fft_features(signal, fs):
    if len(signal) < 8:
        return {
            'total_power': 0.0,
            'spectral_centroid': 0.0,
            **{f'band_{lo}_{hi}Hz': 0.0 for lo, hi in bands}
        }

    signal = detrend(signal)
    N = len(signal)
    freqs = fftfreq(N, d=1/fs)
    power = np.abs(fft(signal))**2

    # Keep only positive frequencies <= 10 Hz
    mask = (freqs > 0) & (freqs <= 10)
    freqs, power = freqs[mask], power[mask]

    total_power = np.sum(power)
    centroid = np.sum(freqs * power) / total_power if total_power > 0 else 0

    features = {
        'total_power': total_power,
        'spectral_centroid': centroid
    }

    for lo, hi in bands:
        features[f'band_{lo}_{hi}Hz'] = np.sum(power[(freqs >= lo) & (freqs < hi)])

    return features

def process_fft_file(input_path):
    try:
        df = pd.read_excel(input_path, sheet_name=0)
        df['datetime'] = pd.to_datetime(df['datetime'], errors='coerce')
        df.dropna(subset=['datetime'], inplace=True)
        df.sort_values('datetime', inplace=True)

        df['interval'] = df['datetime'].dt.floor('10s')

        time_deltas = df['datetime'].diff().dt.total_seconds().dropna()
        fs = 1 / time_deltas.mean() if not time_deltas.empty else 100.0

        records = []
        for interval_time, group in df.groupby('interval'):
            row = {'datetime': interval_time}
            for axis in axes:
                if axis in group.columns:
                    stats = fft_features(group[axis].dropna().values, fs)
                    row.update({f'{axis}_{k}': v for k, v in stats.items()})
            records.append(row)

        fft_df = pd.DataFrame(records)

        # Append to the original file in a new sheet without deleting other sheets
        wb = load_workbook(input_path)
        sheet_name = "FFT_Features"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(title=sheet_name)

        for r in dataframe_to_rows(fft_df, index=False, header=True):
            ws.append(r)

        wb.save(input_path)
        print(f"✅ Embedded FFT features into: {os.path.basename(input_path)}")

    except Exception as e:
        print(f"❌ Error processing {input_path}: {e}")

def recursive_fft_analysis(root_folder):
    for dirpath, _, filenames in os.walk(root_folder):
        for file in filenames:
            if file.endswith('.xlsx') and not file.endswith('_fft_features_cleaned_10s.xlsx'):
                full_path = os.path.join(dirpath, file)
                process_fft_file(full_path)

# === USAGE ===
if __name__ == "__main__":
    root_dir = r"D:\extracted data from JSON file ISI\FINAL BIG DATA FFT"
    recursive_fft_analysis(root_dir)





