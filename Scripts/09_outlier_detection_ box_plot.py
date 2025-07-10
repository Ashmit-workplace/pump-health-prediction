import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def detect_boxplot_outliers(df, axis):
    Q1 = df[axis].quantile(0.25)
    Q3 = df[axis].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    flags = ((df[axis] < lower_bound) | (df[axis] > upper_bound)).astype(int)
    return flags, lower_bound, upper_bound

def add_axiswise_and_combined_flags(df, flag_df):
    # Map flags from flag_df to main df
    for axis in ['x_mps2', 'y_mps2', 'z_mps2']:
        short = axis[0]
        flag_col = f"{axis}_box_flag"
        if flag_col in flag_df.columns:
            new_col = f"{short}_outlier_box_plot"
            df[new_col] = flag_df[flag_col].values

    # Combine all three axis-specific flags into a single column
    combined = df[[f"{a}_outlier_box_plot" for a in ['x', 'y', 'z']] if f"x_outlier_box_plot" in df.columns else []]
    if not combined.empty:
        df['is_outlier_boxplot'] = combined.max(axis=1)
    else:
        df['is_outlier_boxplot'] = 0

    return df

def update_main_sheet_with_flags(filepath, df):
    wb = load_workbook(filepath)
    sheetnames = wb.sheetnames
    main_sheet = sheetnames[0]

    # Remove old main sheet
    wb.remove(wb[main_sheet])
    ws = wb.create_sheet(main_sheet, 0)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(filepath)
    print(f"🟢 Updated main sheet with axis-wise + combined outlier flags in: {os.path.basename(filepath)}")

def process_file_boxplot(filepath):
    try:
        print(f"📄 Processing: {filepath}")
        df = pd.read_excel(filepath)

        if 'datetime' not in df.columns:
            print(f"⚠️ Skipped: 'datetime' column not found.")
            return

        axis_cols = ['x_mps2', 'y_mps2', 'z_mps2']
        outlier_report = []
        flag_df = pd.DataFrame({'datetime': df['datetime']})

        for axis in axis_cols:
            if axis in df.columns:
                flags, lower, upper = detect_boxplot_outliers(df, axis)
                flag_col = f'{axis}_box_flag'
                flag_df[flag_col] = flags

                outliers = df[flags == 1][['datetime', axis]].copy()
                outliers['Axis'] = axis
                outliers['Serial_No'] = outliers.index + 2
                outliers.rename(columns={axis: 'Outlier_Value'}, inplace=True)
                outlier_report.append(outliers)

        # Save flag sheet and report
        if outlier_report:
            report_df = pd.concat(outlier_report, ignore_index=True)
            report_df = report_df[['Serial_No', 'datetime', 'Axis', 'Outlier_Value']]
        else:
            report_df = pd.DataFrame(columns=['Serial_No', 'datetime', 'Axis', 'Outlier_Value'])

        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            flag_df.to_excel(writer, sheet_name="BoxPlot_Flags", index=False)
            report_df.to_excel(writer, sheet_name="BoxPlot_Report", index=False)

        # Add axis-wise and combined flags to main sheet
        df = add_axiswise_and_combined_flags(df, flag_df)
        update_main_sheet_with_flags(filepath, df)

        print(f"✅ Done: Flags, report, and main sheet updated ➤ {os.path.basename(filepath)}\n")

    except Exception as e:
        print(f"❌ Error processing {filepath}: {e}")

def recursive_boxplot_analysis(folder_path):
    for dirpath, _, filenames in os.walk(folder_path):
        for file in filenames:
            if file.endswith('.xlsx'):
                full_path = os.path.join(dirpath, file)
                process_file_boxplot(full_path)

# === USAGE ===
folder_path = r"D:\extracted data from JSON file ISI\rerport writing data\reccurence"
recursive_boxplot_analysis(folder_path)



