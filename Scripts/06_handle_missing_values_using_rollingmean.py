
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === CONFIGURATION ===
ROLLING_WINDOW = 6  # 3 before + 3 after
AXES = ['x', 'y', 'z']
MPS2_AXES = ['x_mps2', 'y_mps2', 'z_mps2']

def impute_missing_with_rolling_mean(df):
    for axis_raw, axis_mps2 in zip(AXES, MPS2_AXES):
        imputed_col = f'{axis_raw}_imputed'
        imputed_vals = []

        for i in range(len(df)):
            if df.at[i, 'is_missing'] == 1 and df.at[i, axis_raw] == 0:
                # Rolling window (3 before, 3 after)
                start = max(0, i - 3)
                end = min(len(df), i + 4)
                window = df.loc[start:end-1, axis_mps2]
                valid_vals = window[window != 0]

                if len(valid_vals) >= 1:
                    imputed_vals.append(valid_vals.mean())
                else:
                    imputed_vals.append('N.A.')
            else:
                imputed_vals.append('N.A.')

        df[imputed_col] = imputed_vals
    return df

def update_excel_safely(filepath):
    try:
        wb = load_workbook(filepath)
        sheetnames = wb.sheetnames
        main_sheet = wb[sheetnames[0]]  # Assuming first sheet is main data sheet

        # Load data from the main sheet into pandas
        df = pd.read_excel(filepath, sheet_name=sheetnames[0])

        if not all(col in df.columns for col in MPS2_AXES + AXES + ['is_missing']):
            print(f"[!] Skipping {filepath} due to missing required columns.")
            return

        df = impute_missing_with_rolling_mean(df)

        # Remove the existing sheet and re-add updated one
        wb.remove(main_sheet)
        ws = wb.create_sheet(sheetnames[0], 0)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        wb.save(filepath)
        print(f"[✓] Imputed columns added safely to: {filepath}")

    except Exception as e:
        print(f"[✗] Error processing {filepath}: {e}")

def process_folder(root_folder):
    for dirpath, _, filenames in os.walk(root_folder):
        for file in filenames:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                full_path = os.path.join(dirpath, file)
                update_excel_safely(full_path)

# === USAGE ===
if __name__ == '__main__':
    folder_path = r"D:\extracted data from JSON file ISI\FINAL BIG DATA\sensor_data"
    process_folder(folder_path)




