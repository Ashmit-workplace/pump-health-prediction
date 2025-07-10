import os
import pandas as pd
import numpy as np
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.cluster import DBSCAN

# === CONFIG ===
EPS_SECONDS = 5
MIN_SAMPLES = 3

Z_SCORE_FLAGS = ['x_outlier_z_score', 'y_outlier_z_score', 'z_outlier_z_score']
BOX_PLOT_FLAGS = ['x_outlier_box_plot', 'y_outlier_box_plot', 'z_outlier_box_plot']

def perform_temporal_clustering(df):
    time_seconds = (df['datetime'] - df['datetime'].min()).dt.total_seconds().values.reshape(-1, 1)
    clustering = DBSCAN(eps=EPS_SECONDS, min_samples=MIN_SAMPLES).fit(time_seconds)
    df['temporal_cluster'] = clustering.labels_
    df['temporal_outlier_type'] = df['temporal_cluster'].apply(lambda x: 'Grouped' if x != -1 else 'Isolated')
    return df

def generate_cluster_report(df):
    report = []
    grouped = df[df['temporal_cluster'] != -1]
    if grouped.empty:
        return pd.DataFrame(columns=['Cluster_ID', 'Count', 'Start_Time', 'End_Time'])
    for cluster_id in sorted(grouped['temporal_cluster'].unique()):
        sub_df = grouped[grouped['temporal_cluster'] == cluster_id]
        report.append({
            'Cluster_ID': cluster_id,
            'Count': len(sub_df),
            'Start_Time': sub_df['datetime'].min(),
            'End_Time': sub_df['datetime'].max()
        })
    return pd.DataFrame(report)

def update_excel_with_temporal_info(filepath):
    try:
        print(f"🕒 Temporal Clustering: {os.path.basename(filepath)}")
        df = pd.read_excel(filepath, sheet_name=0)
        if 'datetime' not in df.columns or ('is_outlier' not in df.columns and 'is_outlier_boxplot' not in df.columns):
            print(f"[!] Skipped: Missing required columns.")
            return
        df['datetime'] = pd.to_datetime(df['datetime'])

        outlier_mask = (df['is_outlier'] == 1) | (df['is_outlier_boxplot'] == 1)
        if not outlier_mask.any():
            print("   ⚠ No outliers found for clustering.")
            return

        outlier_df = df[outlier_mask].copy()
        outlier_df['original_index'] = outlier_df.index  # Store original index

        outlier_df = perform_temporal_clustering(outlier_df)

        # Initialize all as NaN/Normal
        df['temporal_cluster'] = np.nan
        df['temporal_outlier_type'] = "Normal"

        # Restore the clustered values
        df.loc[outlier_df['original_index'], 'temporal_cluster'] = outlier_df['temporal_cluster'].values
        df.loc[outlier_df['original_index'], 'temporal_outlier_type'] = outlier_df['temporal_outlier_type'].values

        # Generate cluster report
        cluster_report = generate_cluster_report(outlier_df)

        # Save updates to the Excel file
        wb = load_workbook(filepath)
        main_sheet = wb.sheetnames[0]
        wb.remove(wb[main_sheet])
        ws = wb.create_sheet(main_sheet, 0)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            cluster_report.to_excel(writer, sheet_name="Temporal_Cluster_Report", index=False)

        wb.save(filepath)
        print(f"✅ Saved: Temporal clustering info added to {os.path.basename(filepath)}")

    except Exception as e:
        print(f"[✗] Error in {filepath}: {e}")

def process_folder_temporal_clustering(root_folder):
    for dirpath, _, filenames in os.walk(root_folder):
        for file in filenames:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                full_path = os.path.join(dirpath, file)
                update_excel_with_temporal_info(full_path)

# === USAGE ===
if __name__ == "__main__":
    folder_path = r"D:\extracted data from JSON file ISI\rerport writing data\reccurence"
    process_folder_temporal_clustering(folder_path)



