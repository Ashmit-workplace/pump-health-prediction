import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === CONFIGURATION ===
ROLLING_WINDOW = 6  # 3 before + 3 after
AXES = ['x', 'y', 'z']
MPS2_AXES = ['x_mps2', 'y_mps2', 'z_mps2']
Z_SCORE_FLAGS = ['x_outlier_z_score', 'y_outlier_z_score', 'z_outlier_z_score']
BOX_PLOT_FLAGS = ['x_outlier_box_plot', 'y_outlier_box_plot', 'z_outlier_box_plot']

def impute_outlier_with_rolling_mean(df):
    for axis_raw, axis_mps2, z_flag, box_flag in zip(AXES, MPS2_AXES, Z_SCORE_FLAGS, BOX_PLOT_FLAGS):
        imputed_col = f'{axis_raw}_imputed'
        imputed_vals = []

        for i in range(len(df)):
            is_z_outlier = df.at[i, z_flag] if z_flag in df.columns else 0
            is_box_outlier = df.at[i, box_flag] if box_flag in df.columns else 0

            if is_z_outlier == 1 or is_box_outlier == 1:
                start = max(0, i - 3)
                end = min(len(df), i + 4)
                window = df.loc[start:end-1, axis_mps2]
                valid_vals = window[window != 0]

                if len(valid_vals) >= 1:
                    imputed_vals.append(valid_vals.mean())
                else:
                    imputed_vals.append('N.A.')
            else:
                imputed_vals.append('N.A.')

        df[imputed_col] = imputed_vals
    return df

def update_excel_safely(filepath):
    try:
        wb = load_workbook(filepath)
        sheetnames = wb.sheetnames
        main_sheet = wb[sheetnames[0]]

        # Load main sheet as dataframe
        df = pd.read_excel(filepath, sheet_name=sheetnames[0])

        required_cols = MPS2_AXES + AXES + Z_SCORE_FLAGS + BOX_PLOT_FLAGS
        if not all(col in df.columns for col in required_cols):
            print(f"[!] Skipping {filepath} due to missing required columns.")
            return

        df = impute_outlier_with_rolling_mean(df)

        # Overwrite main sheet only
        wb.remove(main_sheet)
        ws = wb.create_sheet(sheetnames[0], 0)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        wb.save(filepath)
        print(f"[✓] Imputed values added in: {os.path.basename(filepath)}")

    except Exception as e:
        print(f"[✗] Error updating file: {filepath}\n    └─▶ {e}")

def process_folder(root_folder):
    for dirpath, _, filenames in os.walk(root_folder):
        for file in filenames:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                full_path = os.path.join(dirpath, file)
                update_excel_safely(full_path)

# === USAGE ===
if __name__ == '__main__':
    folder_path = r"D:\extracted data from JSON file ISI\extracted data\sensor_data_CLEANED\z_score analysis for outliers"
    process_folder(folder_path)









