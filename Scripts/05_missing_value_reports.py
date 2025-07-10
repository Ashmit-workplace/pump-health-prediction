import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === CONFIGURATION ===
ROLLING_WINDOW = 6
UNRELIABLE_THRESHOLD = 3  # Threshold for number of missing points per window


def create_imputation_plot_data(df):
    imputed_points = df[df['is_missing'] == 1].copy()
    if imputed_points.empty:
        return pd.DataFrame()

    combined_rows = []
    for idx in imputed_points.index:
        start = max(0, idx - 3)
        end = min(len(df), idx + 4)
        local = df.iloc[start:end].copy()
        local['x_used'] = np.where(local['x_mps2'] == 0, local['x_imputed'], local['x_mps2'])
        local['y_used'] = np.where(local['y_mps2'] == 0, local['y_imputed'], local['y_mps2'])
        local['z_used'] = np.where(local['z_mps2'] == 0, local['z_imputed'], local['z_mps2'])
        combined_rows.append(local[['datetime', 'x_mps2', 'x_used', 'y_mps2', 'y_used', 'z_mps2', 'z_used']])

    result = pd.concat(combined_rows).drop_duplicates().reset_index(drop=True)
    return result


def create_missingness_summary(df):
    df['second'] = df['datetime'].dt.floor('1s')
    summary = df.groupby('second')['is_missing'].sum().reset_index()
    summary.columns = ['Timestamp_Second', 'Missing_Count']
    return summary


def create_unreliable_windows(df):
    df['window'] = df['datetime'].dt.floor('10s')
    windows = df.groupby('window')['is_missing'].sum().reset_index()
    windows = windows[windows['is_missing'] > UNRELIABLE_THRESHOLD].copy()
    windows['End_Time'] = windows['window'] + pd.Timedelta(seconds=10)
    return windows[['window', 'End_Time', 'is_missing']].rename(columns={'window': 'Start_Time', 'is_missing': 'Missing_Count'})


def write_analysis_to_excel(filepath):
    try:
        df = pd.read_excel(filepath)
        df['datetime'] = pd.to_datetime(df['datetime'], errors='coerce')
        df.dropna(subset=['datetime'], inplace=True)

        wb = load_workbook(filepath)

        # === Imputation Plot View Sheet ===
        plot_df = create_imputation_plot_data(df)
        if not plot_df.empty:
            ws1 = wb.create_sheet("Imputation_Plot_View")
            for r in dataframe_to_rows(plot_df, index=False, header=True):
                ws1.append(r)

        # === Missingness Summary ===
        summary_df = create_missingness_summary(df)
        ws2 = wb.create_sheet("Missingness_Pattern")
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws2.append(r)

        # === Unreliable Windows ===
        unreliable_df = create_unreliable_windows(df)
        ws3 = wb.create_sheet("Unreliable_Windows")
        for r in dataframe_to_rows(unreliable_df, index=False, header=True):
            ws3.append(r)

        wb.save(filepath)
        print(f"✅ Updated: {filepath}")

    except Exception as e:
        print(f"❌ Error processing {filepath}: {e}")


def recursive_add_analysis(folder_path):
    for dirpath, _, filenames in os.walk(folder_path):
        for file in filenames:
            if file.endswith("_flagged_missing.xlsx"):
                full_path = os.path.join(dirpath, file)
                write_analysis_to_excel(full_path)


# === USAGE ===
root_folder = r"D:\extracted data from JSON file ISI\extracted data\sensor_data_CLEANED\sensor_data_cleaned_original"
recursive_add_analysis(root_folder)




