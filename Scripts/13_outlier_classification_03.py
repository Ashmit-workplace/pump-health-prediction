import os
import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === CONFIGURATION ===
SEGMENT_DURATION = 15  # seconds
OFFSET_TOLERANCE = 0.5  # seconds for recurrence matching
MIN_RECURSIONS = 3  # how many segments must repeat the same offset

def detect_recurring_offsets(df):
    df['datetime'] = pd.to_datetime(df['datetime'])
    df['time_offset'] = (df['datetime'] - df['datetime'].min()).dt.total_seconds()
    df['segment_id'] = (df['time_offset'] // SEGMENT_DURATION).astype(int)
    df['offset_in_segment'] = df['time_offset'] % SEGMENT_DURATION

    # Only consider outliers
    outlier_df = df[df['is_outlier'] == 1].copy()

    offset_map = defaultdict(set)
    for _, row in outlier_df.iterrows():
        rounded_offset = round(row['offset_in_segment'] / OFFSET_TOLERANCE) * OFFSET_TOLERANCE
        offset_map[rounded_offset].add(row['segment_id'])

    recurring_offsets = {
        offset for offset, segments in offset_map.items() if len(segments) >= MIN_RECURSIONS
    }

    df['recurring_anomaly'] = df.apply(
       lambda row: int(
           row['is_outlier'] == 1 and
           round(row['offset_in_segment'] / OFFSET_TOLERANCE) * OFFSET_TOLERANCE in recurring_offsets
        ), axis=1
    )

    # Add recurrence strength if needed
    recurrence_strength = {
        offset: len(segments) for offset, segments in offset_map.items()
    }

    df['recurrence_score'] = df.apply(
        lambda row: recurrence_strength.get(
            round(row['offset_in_segment'] / OFFSET_TOLERANCE) * OFFSET_TOLERANCE, 0
        ) if row['is_outlier'] == 1 else 0, axis=1
    )

    return df

def analyze_file_recurrence(filepath):
    print(f"🔁 Processing Recurrence: {os.path.basename(filepath)}")
    try:
        df = pd.read_excel(filepath)
        if 'datetime' not in df.columns or 'is_outlier' not in df.columns:
            print("⚠️ Missing required columns.")
            return

        df = detect_recurring_offsets(df)

        # Overwrite only the main sheet
        wb = load_workbook(filepath)
        sheetnames = wb.sheetnames
        main_sheet = sheetnames[0]
        wb.remove(wb[main_sheet])
        ws = wb.create_sheet(main_sheet, 0)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        wb.save(filepath)
        print(f"✅ Saved: Recurrence results added ➤ {os.path.basename(filepath)}")

    except Exception as e:
        print(f"[✗] Error processing {filepath}: {e}")

def process_folder(root_dir):
    for dirpath, _, filenames in os.walk(root_dir):
        for file in filenames:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                analyze_file_recurrence(os.path.join(dirpath, file))

# === MAIN EXECUTION ===
if __name__ == "__main__":
    process_folder(r"D:\extracted data from JSON file ISI\rerport writing data\reccurence")
