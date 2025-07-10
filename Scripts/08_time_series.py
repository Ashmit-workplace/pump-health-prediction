import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === CONFIGURATION ===
WINDOW_SIZE = 51
RMS_STD_MULTIPLIER = 2.0
KURTOSIS_FIXED_THRESHOLD = 3.5
PERCENTILE = 0.95
AXES = ['x', 'y', 'z']
INPUT_COLUMNS = [f"{axis}_mps2" for axis in AXES]

def generate_combined_flag_report(df):
    """Generate a report for all RMS and kurtosis flagged rows."""
    records = []
    for axis in AXES:
        for flag_type in ['kurt_fixed_flag', 'kurt_percentile_flag', 'rms_fixed_flag', 'rms_percentile_flag']:
            flag_col = f"{flag_type}_{axis}"
            if flag_col in df.columns:
                flagged = df[df[flag_col] == True]
                for idx in flagged.index:
                    timestamp = flagged.at[idx, 'datetime'] if 'datetime' in flagged.columns else flagged.at[idx, 'timestamp']
                    records.append({
                        'Serial_No': idx + 2,
                        'Timestamp': timestamp,
                        'Criteria': flag_col,
                        'Axis': axis
                    })
    return pd.DataFrame(records)

def process_file_inplace(filepath):
    try:
        df = pd.read_excel(filepath)

        if not all(col in df.columns for col in INPUT_COLUMNS):
            print(f"[!] Skipping {os.path.basename(filepath)}: Missing required columns.")
            return

        for axis in AXES:
            col = f"{axis}_mps2"

            # === Rolling RMS with strict window ===
            rms_col = f"rolling_rms_{axis}"
            rolling_rms = df[col].rolling(WINDOW_SIZE, center=True, min_periods=WINDOW_SIZE)\
                                  .apply(lambda s: np.sqrt(np.mean(s**2)))
            df[rms_col] = rolling_rms.fillna(0)  # Drop edge values by replacing with 0

            # === Rolling Kurtosis with strict window ===
            kurt_col = f"rolling_kurtosis_{axis}"
            rolling_kurt = df[col].rolling(WINDOW_SIZE, center=True, min_periods=WINDOW_SIZE).kurt()
            df[kurt_col] = rolling_kurt.fillna(0)  # Drop edge values by replacing with 0

            # === RMS Thresholding ===
            rms_mean = df[rms_col].mean()
            rms_std = df[rms_col].std()
            fixed_rms_thresh = rms_mean + RMS_STD_MULTIPLIER * rms_std
            perc_rms_thresh = df[rms_col].quantile(PERCENTILE)

            df[f"rms_fixed_flag_{axis}"] = df[rms_col] > fixed_rms_thresh
            df[f"rms_percentile_flag_{axis}"] = df[rms_col] > perc_rms_thresh
            df[f"rms_combined_flag_{axis}"] = df[f"rms_fixed_flag_{axis}"] | df[f"rms_percentile_flag_{axis}"]

            # === Kurtosis Thresholding ===
            fixed_kurt_thresh = KURTOSIS_FIXED_THRESHOLD
            perc_kurt_thresh = df[kurt_col].quantile(PERCENTILE)

            df[f"kurt_fixed_flag_{axis}"] = df[kurt_col] > fixed_kurt_thresh
            df[f"kurt_percentile_flag_{axis}"] = df[kurt_col] > perc_kurt_thresh
            df[f"kurt_combined_flag_{axis}"] = df[f"kurt_fixed_flag_{axis}"] | df[f"kurt_percentile_flag_{axis}"]

        # === Safe Overwrite of Main Sheet ===
        wb = load_workbook(filepath)
        main_sheet = wb.sheetnames[0]
        wb.remove(wb[main_sheet])
        ws = wb.create_sheet(main_sheet, 0)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # === Add Flag Report Sheet ===
        report_df = generate_combined_flag_report(df)
        if not report_df.empty:
            ws_report = wb.create_sheet("RollingStats_Report")
            for r in dataframe_to_rows(report_df, index=False, header=True):
                ws_report.append(r)

        wb.save(filepath)
        print(f"[✓] Updated with RollingStats_Report: {os.path.basename(filepath)}")

    except Exception as e:
        print(f"[✗] Error processing {filepath}: {e}")

def process_folder_recursive_inplace(input_root):
    file_count = 0
    for dirpath, _, filenames in os.walk(input_root):
        for file in filenames:
            if file.endswith('.xlsx') and not file.startswith('~$'):
                full_path = os.path.join(dirpath, file)
                process_file_inplace(full_path)
                file_count += 1
    print(f"\n[✓] Finished updating {file_count} file(s).")

# === USAGE ===
if __name__ == "__main__":
    input_root_folder = r"D:\extracted data from JSON file ISI\FINAL BIG DATA\sensor_data"
    process_folder_recursive_inplace(input_root_folder)




