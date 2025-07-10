import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from PIL import Image


def analyze_spikes_and_embed(filepath, std_dev_threshold=3.0, use_adaptive_threshold=True, quantile_threshold=0.99):
    try:
        df = pd.read_excel(filepath)

        # Parse datetime
        if 'timestamp' in df.columns:
            df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
        elif 'datetime' in df.columns:
            df['datetime'] = pd.to_datetime(df['datetime'])
        else:
            print(f"[ERROR] No 'timestamp' or 'datetime' column found in: {filepath}")
            return

        # Detect axes
        if all(col in df.columns for col in ['x', 'y', 'z']):
            axes = ['x', 'y', 'z']
        elif all(col in df.columns for col in ['x_mps2', 'y_mps2', 'z_mps2']):
            axes = ['x_mps2', 'y_mps2', 'z_mps2']
        else:
            print(f"[ERROR] Required axis columns not found in: {filepath}")
            return

        print(f"\n📊 Analyzing: {os.path.basename(filepath)}")

        all_spikes = []
        summary_stats = []
        peak_points = []

        for axis in axes:
            df[axis] = df[axis].ffill().bfill()
            mean = df[axis].mean()
            std = df[axis].std()
            z_col = f"{axis}_zscore"
            df[z_col] = (df[axis] - mean) / std if std > 0 else 0

            # Adaptive thresholds based on quantiles
            upper_thresh = df[axis].quantile(quantile_threshold)
            lower_thresh = df[axis].quantile(1 - quantile_threshold)

            outlier_flag_col = f"{axis[0]}_outlier_z_score"
            if use_adaptive_threshold:
                df[outlier_flag_col] = (df[axis] > upper_thresh) | (df[axis] < lower_thresh)
            else:
                df[outlier_flag_col] = df[z_col].abs() > std_dev_threshold

            spikes = df[df[outlier_flag_col]]
            if not spikes.empty:
                spike_info = spikes[['datetime', axis, z_col]].copy()
                spike_info['Serial_No'] = spikes.index + 2
                spike_info['Axis'] = axis
                spike_info.rename(columns={axis: 'Spike_Value', z_col: 'Z_Score'}, inplace=True)
                all_spikes.append(spike_info)

                max_idx = spike_info['Z_Score'].abs().idxmax()
                peak_row = spike_info.loc[max_idx]
                peak_points.append(peak_row)

            summary_stats.append({
                "Axis": axis,
                "Mean": mean,
                "Std Dev": std,
                "Upper Threshold": upper_thresh,
                "Lower Threshold": lower_thresh,
                "Max Z-score": df[z_col].abs().max(),
                "Spikes Found": len(spikes),
                "% Spikes": round(len(spikes) / len(df) * 100, 2)
            })

        # Combine axis-specific flags into a single is_outlier column
        axis_flags = [f"{axis[0]}_outlier_z_score" for axis in axes]
        df['is_outlier'] = df[axis_flags].any(axis=1).astype(int)

        # Generate plots
        fig1, axs1 = plt.subplots(len(axes), 1, figsize=(12, 8), sharex=True)
        for i, axis in enumerate(axes):
            flag_col = f"{axis[0]}_outlier_z_score"
            axs1[i].plot(df['datetime'], df[axis], label=f'{axis} signal')
            axs1[i].scatter(df.loc[df[flag_col], 'datetime'], df.loc[df[flag_col], axis], color='red', label='Spikes')
            axs1[i].legend()
            axs1[i].set_ylabel("Amplitude")
            axs1[i].grid(True)
        axs1[-1].set_xlabel("Time")
        fig1.tight_layout()
        buf1 = BytesIO()
        fig1.savefig(buf1, format='png')
        plt.close(fig1)

        fig2, axs2 = plt.subplots(len(axes), 1, figsize=(12, 8), sharex=True)
        for i, axis in enumerate(axes):
            z = df[f"{axis}_zscore"]
            axs2[i].plot(df['datetime'], z, label=f'{axis} Z-score', color='green')
            axs2[i].axhline(std_dev_threshold, color='red', linestyle='--', label='±Threshold')
            axs2[i].axhline(-std_dev_threshold, color='red', linestyle='--')
            axs2[i].legend()
            axs2[i].set_ylabel("Z-Score")
            axs2[i].grid(True)
        axs2[-1].set_xlabel("Time")
        fig2.tight_layout()
        buf2 = BytesIO()
        fig2.savefig(buf2, format='png')
        plt.close(fig2)

        # === Update Excel File ===
        wb = load_workbook(filepath)
        main_sheet = wb.sheetnames[0]
        wb.remove(wb[main_sheet])
        ws = wb.create_sheet(main_sheet, 0)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Add or replace Plots sheet
        if 'Plots' in wb.sheetnames:
            wb.remove(wb['Plots'])
        ws_plot = wb.create_sheet("Plots")

        # Convert plots to Pillow image and insert
        for i, buf in enumerate([buf1, buf2]):
            buf.seek(0)
            img = Image.open(buf)
            image_path = f"temp_plot_{i}.png"
            img.save(image_path)
            xl_img = XLImage(image_path)
            ws_plot.add_image(xl_img, f"B{2 + i * 30}")

        # Spike Report
        if all_spikes:
            spike_df = pd.concat(all_spikes, ignore_index=True)
            spike_df['datetime'] = pd.to_datetime(spike_df['datetime']).dt.strftime('%Y-%m-%d %H:%M:%S.%f').str[:-3]
            if 'Spike_Report' in wb.sheetnames:
                wb.remove(wb['Spike_Report'])
            ws_report = wb.create_sheet("Spike_Report")
            for r in dataframe_to_rows(spike_df, index=False, header=True):
                ws_report.append(r)

        # Summary Stats
        if 'Summary_Stats' in wb.sheetnames:
            wb.remove(wb['Summary_Stats'])
        ws_summary = wb.create_sheet("Summary_Stats")
        for r in dataframe_to_rows(pd.DataFrame(summary_stats), index=False, header=True):
            ws_summary.append(r)

        # Peak Points
        if peak_points:
            peak_df = pd.DataFrame(peak_points)
            peak_df['datetime'] = pd.to_datetime(peak_df['datetime']).dt.strftime('%Y-%m-%d %H:%M:%S.%f').str[:-3]
            peak_df = peak_df[['Serial_No', 'datetime', 'Axis', 'Spike_Value', 'Z_Score']]
            if 'Peak_Spike_Coordinates' in wb.sheetnames:
                wb.remove(wb['Peak_Spike_Coordinates'])
            ws_peaks = wb.create_sheet("Peak_Spike_Coordinates")
            for r in dataframe_to_rows(peak_df, index=False, header=True):
                ws_peaks.append(r)

        wb.save(filepath)
        print(f"[✓] Embedded and saved to: {os.path.basename(filepath)}")

    except Exception as e:
        print(f"[✗] Error processing {filepath}: {e}")


def recursive_spike_analysis(root_folder):
    for dirpath, _, filenames in os.walk(root_folder):
        for file in filenames:
            if file.endswith('.xlsx') and not file.endswith('_analysis_report_embedded.xlsx'):
                full_path = os.path.join(dirpath, file)
                analyze_spikes_and_embed(full_path)


if __name__ == "__main__":
    folder_path = r"D:\extracted data from JSON file ISI\rerport writing data\reccurence"
    recursive_spike_analysis(folder_path)










