import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# === CONFIGURATION ===
bands = [(0, 1), (1, 3), (3, 5), (5, 10)]  # Frequency bands up to 10 Hz
axes = ['x', 'y', 'z']

def normalize_columns(df, cols):
    df = df.copy()
    for col in cols:
        if df[col].dtype in [float, int]:
            min_val, max_val = df[col].min(), df[col].max()
            if max_val != min_val:
                df[col] = (df[col] - min_val) / (max_val - min_val)
            else:
                df[col] = 0
    return df

def process_excel_file(filepath):
    try:
        print(f"Processing: {filepath}")

        # Load main and FFT sheets
        df_main = pd.read_excel(filepath, sheet_name=0)
        df_fft = pd.read_excel(filepath, sheet_name='FFT_Features')

        # Preprocessing
        df_main['datetime'] = pd.to_datetime(df_main['datetime'], errors='coerce')
        df_fft['interval'] = pd.to_datetime(df_fft['datetime'], errors='coerce')
        
        # ✅ FIXED INTERVAL ALIGNMENT (removed +10s shift)
        df_main['interval'] = df_main['datetime'].dt.floor('10s')

        # --- Time-Domain Score with Weights ---
        for axis in axes:
            df_main[f'rms_combined_flag_{axis}'] = df_main[f'rms_combined_flag_{axis}'].astype(int)
            df_main[f'kurt_combined_flag_{axis}'] = df_main[f'kurt_combined_flag_{axis}'].astype(int)

        df_main['rms_score'] = df_main[[f'rms_combined_flag_{a}' for a in axes]].sum(axis=1) / 3
        df_main['kurt_score'] = df_main[[f'kurt_combined_flag_{a}' for a in axes]].sum(axis=1) / 3
        df_main['time_series_score'] = (df_main['rms_score'] + df_main['kurt_score']) / 2

        df_main['contextual_score'] = df_main['final_contextual_score']
        df_main['temporal_score'] = df_main['temporal_outlier_type'].apply(lambda x: 1 if x == 'Grouped' else 0)
        max_rec_score = df_main['recurrence_score'].max()
        df_main['recurrence_score'] = (
            df_main['recurrence_score'] / max_rec_score if max_rec_score > 0 else 0
        )

        df_main['time_domain_score'] = (
            0.5 * df_main['time_series_score'] +
            0.2 * df_main['contextual_score'] +
            0.2 * df_main['temporal_score'] +
            0.1 * df_main['recurrence_score']
        )

        # --- Frequency Score ---
        fft_features_cols = []
        for axis in axes:
            prefix = f"{axis}_mps2"
            fft_features_cols += [
                f'{prefix}_total_power',
                f'{prefix}_spectral_centroid',
                f'{prefix}_band_0_1Hz',
                f'{prefix}_band_1_3Hz',
                f'{prefix}_band_3_5Hz',
                f'{prefix}_band_5_10Hz']

        df_fft = normalize_columns(df_fft, fft_features_cols)

        for axis in axes:
            prefix = f"{axis}_mps2"
            df_fft[f'{axis}_score'] = df_fft[[ 
                f'{prefix}_total_power',
                f'{prefix}_spectral_centroid',
                f'{prefix}_band_0_1Hz',
                f'{prefix}_band_1_3Hz',
                f'{prefix}_band_3_5Hz',
                f'{prefix}_band_5_10Hz']].mean(axis=1)

        df_fft['frequency_interval_score'] = df_fft[[f'{a}_score' for a in axes]].mean(axis=1)

        # --- Merge Scores ---
        df_main = df_main.merge(df_fft[['interval', 'frequency_interval_score']], on='interval', how='left')
        df_main.rename(columns={'frequency_interval_score': 'time_based_frequency_score'}, inplace=True)
        df_main['time_based_frequency_score'] = df_main['time_based_frequency_score'].fillna(0)

        df_main['Final_score'] = (df_main['time_domain_score'] + df_main['time_based_frequency_score']) / 2

        # --- Custom Quantile-Based Labeling ---
        q2 = df_main['Final_score'].quantile(0.50)
        q3 = df_main['Final_score'].quantile(0.75)
        q99 = df_main['Final_score'].quantile(0.95)

        def label_row(score):
            if score > q99:
                return 'Critical'
            elif score > q3:
                return 'Warning'
            elif score > q2:
                return 'Monitor'
            else:
                return 'Healthy'

        df_main['Final_label'] = df_main['Final_score'].apply(label_row)

        # === Excel Writing ===
        wb = load_workbook(filepath)
        main_sheet = wb.sheetnames[0]
        wb.remove(wb[main_sheet])
        ws = wb.create_sheet(main_sheet, 0)

        # Define label colors
        label_colors = {
            "Healthy": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Green
            "Monitor": PatternFill(start_color="FFFCCC", end_color="FFFCCC", fill_type="solid"),  # Yellow
            "Warning": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # Orange
            "Critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        }

        for i, r in enumerate(dataframe_to_rows(df_main, index=False, header=True)):
            ws.append(r)
            if i == 0:
                label_col_idx = r.index("Final_label") + 1
            else:
                label = r[label_col_idx - 1]
                if label in label_colors:
                    ws.cell(row=i + 1, column=label_col_idx).fill = label_colors[label]

        wb.save(filepath)
        print(f"✅ Done: Scoring and revised labeling updated in {os.path.basename(filepath)}")

    except Exception as e:
        print(f"❌ Error in {filepath}: {e}")

def recursive_scoring_runner(root_folder):
    for dirpath, _, filenames in os.walk(root_folder):
        for file in filenames:
            if file.endswith('.xlsx') and not file.startswith('~$'):
                filepath = os.path.join(dirpath, file)
                process_excel_file(filepath)

# === USAGE ===
if __name__ == '__main__':
    root_dir = r"D:\extracted data from JSON file ISI\FINAL BIG DATA FFT SCORED"
    recursive_scoring_runner(root_dir)





