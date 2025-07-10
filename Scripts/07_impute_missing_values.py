import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def integrate_imputed_values(filepath):
    print(f"📄 Checking: {os.path.basename(filepath)}")
    try:
        df = pd.read_excel(filepath, sheet_name=0, na_values=["N.A.", "NA", "n.a.", "na"])

        updated = False
        for axis in ['x', 'y', 'z']:
            raw_col = f"{axis}_mps2"
            imputed_col = f"{axis}_imputed"
            if raw_col in df.columns and imputed_col in df.columns:
                # Only update where imputed value is not NaN
                non_null_mask = df[imputed_col].notna()
                before = df[raw_col].copy()
                df.loc[non_null_mask, raw_col] = df.loc[non_null_mask, imputed_col]
                if not before.equals(df[raw_col]):
                    updated = True

        if updated:
            wb = load_workbook(filepath)
            main_sheet = wb.sheetnames[0]
            wb.remove(wb[main_sheet])
            ws = wb.create_sheet(main_sheet, 0)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            wb.save(filepath)
            print(f"✅ Imputed values integrated in {os.path.basename(filepath)}")
        else:
            print(f"⏩ Skipped (no imputed values to apply)")

    except Exception as e:
        print(f"[✗] Error in {filepath}: {e}")

def recursive_imputation_integration(folder_path):
    for dirpath, _, filenames in os.walk(folder_path):
        for file in filenames:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                full_path = os.path.join(dirpath, file)
                integrate_imputed_values(full_path)

# === USAGE ===
if __name__ == "__main__":
    folder_path = r"D:\extracted data from JSON file ISI\FINAL BIG DATA\sensor_data"
    recursive_imputation_integration(folder_path)