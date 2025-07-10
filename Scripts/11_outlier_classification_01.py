import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Define flag columns
Z_SCORE_FLAGS = ['x_outlier_z_score', 'y_outlier_z_score', 'z_outlier_z_score']
BOX_PLOT_FLAGS = ['x_outlier_box_plot', 'y_outlier_box_plot', 'z_outlier_box_plot']

# Utility: count True flags
def count_true_flags(row, cols):
    return sum(bool(row.get(col, False)) for col in cols)

# === Logic 1: Loosened Yet Structured ===
def evaluate_contextual_label_loosened(df, idx):
    row = df.iloc[idx]
    curr_z = count_true_flags(row, Z_SCORE_FLAGS)
    curr_box = count_true_flags(row, BOX_PLOT_FLAGS)
    curr_total = curr_z + curr_box

    neighbor_total = 0
    for j in [idx - 1, idx + 1]:
        if 0 <= j < len(df):
            neighbor_total += count_true_flags(df.iloc[j], Z_SCORE_FLAGS + BOX_PLOT_FLAGS)

    is_outlier_flagged = row['is_outlier'] == 1 or row['is_outlier_boxplot'] == 1

    if is_outlier_flagged:
        if curr_total >= 1:
            if neighbor_total >= 1 or curr_total >= 2:
                return "True Anomaly"
            else:
                return "Likely Sensor Fault"
        else:
            return "Uncertain"
    elif neighbor_total >= 2:
        return "Uncertain"
    else:
        return "Normal"

# === Logic 2: Enhanced Detection-Specific ===
def evaluate_contextual_label_enhanced(df, idx):
    row = df.iloc[idx]
    curr_z = count_true_flags(row, Z_SCORE_FLAGS)
    curr_box = count_true_flags(row, BOX_PLOT_FLAGS)
    zscore_flag = curr_z > 0
    boxplot_flag = curr_box > 0

    neighbor_total = 0
    for j in [idx - 1, idx + 1]:
        if 0 <= j < len(df):
            neighbor_total += count_true_flags(df.iloc[j], Z_SCORE_FLAGS + BOX_PLOT_FLAGS)

    if zscore_flag and boxplot_flag:
        if neighbor_total >= 1:
            return "True Anomaly"
        else:
            return "Likely Mechanical Fault (Weak Context)"
    elif zscore_flag or boxplot_flag:
        if zscore_flag and not boxplot_flag and neighbor_total >= 1:
            return "Likely Mechanical Fault (Z-score only)"
        elif neighbor_total >= 2:
            return "Likely Sensor Fault with Context"
        else:
            return "Uncertain"
    elif neighbor_total >= 2:
        return "Suspicious Region"
    else:
        return "Normal"

# Apply both label strategies + summarize
def apply_contextual_labeling_methods(df):
    df['loosened_contextual_label'] = [evaluate_contextual_label_loosened(df, i) for i in range(len(df))]
    df['enhanced_contextual_label'] = [evaluate_contextual_label_enhanced(df, i) for i in range(len(df))]

    label_score_map_loosened = {
        "Normal": 0,
        "Uncertain": 1,
        "Likely Sensor Fault": 2,
        "True Anomaly": 3
    }

    label_score_map_enhanced = {
        "Normal": 0,
        "Suspicious Region": 1,
        "Uncertain": 2,
        "Likely Mechanical Fault (Weak Context)": 2.5,
        "Likely Sensor Fault with Context": 3,
        "Likely Mechanical Fault (Z-score only)": 3.5,
        "True Anomaly": 4
    }

    df['contextual_score_loosened'] = df['loosened_contextual_label'].map(label_score_map_loosened)
    df['contextual_score_enhanced'] = df['enhanced_contextual_label'].map(label_score_map_enhanced)

    df['final_contextual_score'] = (
        0.5 * (df['contextual_score_loosened'] / 3) +
        0.5 * (df['contextual_score_enhanced'] / 4)
    )

    def assign_final_label(score):
        if score < 0.25:
            return "Normal"
        elif score < 0.5:
            return "Mild Anomaly"
        elif score < 0.75:
            return "Probable Fault"
        else:
            return "Confirmed Anomaly"

    df['final_contextual_label'] = df['final_contextual_score'].apply(assign_final_label)
    return df

# Overwrite only main sheet, preserve others
def update_main_sheet_preserving_others(filepath, updated_df):
    wb = load_workbook(filepath)
    main_sheet_name = wb.sheetnames[0]

    # Remove old main sheet
    wb.remove(wb[main_sheet_name])
    ws = wb.create_sheet(main_sheet_name, 0)

    for r in dataframe_to_rows(updated_df, index=False, header=True):
        ws.append(r)

    wb.save(filepath)
    print(f"[✓] Contextual labels safely added to: {os.path.basename(filepath)}")

# File processor function
def process_file(filepath):
    df = pd.read_excel(filepath, sheet_name=0)

    required_cols = Z_SCORE_FLAGS + BOX_PLOT_FLAGS + ['is_outlier', 'is_outlier_boxplot']
    if not all(col in df.columns for col in required_cols):
        print(f"[!] Missing required columns in: {filepath}")
        return

    df = apply_contextual_labeling_methods(df)
    update_main_sheet_preserving_others(filepath, df)

# Folder processor
def process_folder(root_dir):
    for dirpath, _, filenames in os.walk(root_dir):
        for file in filenames:
            if file.endswith('.xlsx'):
                process_file(os.path.join(dirpath, file))

# === USAGE ===
if __name__ == '__main__':
    folder_path = r"D:\extracted data from JSON file ISI\rerport writing data\reccurence"
    process_folder(folder_path)
